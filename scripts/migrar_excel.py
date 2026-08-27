"""
Migración de Control de Gastos.xlsx → Supabase

Uso:
  pip install openpyxl requests python-dateutil
  python scripts/migrar_excel.py --file "Control de Gastos.xlsx"
  python scripts/migrar_excel.py --file "Control de Gastos.xlsx" --dry-run
  python scripts/migrar_excel.py --file "Control de Gastos.xlsx" --show-sheets
"""

import argparse
import sys
from datetime import date, datetime
from dateutil.relativedelta import relativedelta  # pip install python-dateutil
from pathlib import Path
import requests

try:
    import openpyxl
except ImportError:
    sys.exit("Faltan dependencias. Corré: pip install openpyxl requests python-dateutil")

# ─── Config ──────────────────────────────────────────────────

def load_env():
    env_path = Path(__file__).parent.parent / ".env.local"
    env = {}
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env

env = load_env()
SUPABASE_URL = env.get("NEXT_PUBLIC_SUPABASE_URL", "")
SUPABASE_KEY = env.get("NEXT_PUBLIC_SUPABASE_ANON_KEY", "")

if not SUPABASE_URL or not SUPABASE_KEY:
    sys.exit("ERROR: Completá el .env.local con la URL y anon key de Supabase antes de migrar.")

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

def sb_get(table, params=""):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/{table}?{params}", headers=HEADERS)
    r.raise_for_status()
    return r.json()

def sb_post(table, rows):
    if not rows:
        return []
    r = requests.post(f"{SUPABASE_URL}/rest/v1/{table}", headers=HEADERS, json=rows)
    if not r.ok:
        print(f"  ERROR en {table}: {r.text[:400]}")
        return []
    return r.json() if r.text else []

# ─── Helpers ──────────────────────────────────────────────────

def parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None

def parse_float(val) -> float:
    """Siempre retorna float (0.0 si vacío). Usá para campos requeridos."""
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(",", ".").replace("$", "").strip())
    except Exception:
        return 0.0

def parse_float_or_none(val) -> float | None:
    """Retorna None para vacío/cero. Usá para campos opcionales (amount_usd, ccl_rate)."""
    if val is None:
        return None
    s = str(val).replace(",", ".").replace("$", "").strip()
    if not s:
        return None
    try:
        result = float(s)
        return result if result != 0.0 else None
    except Exception:
        return None

def first_day_next_month(d: str) -> str:
    dt = datetime.strptime(d, "%Y-%m-%d")
    next_month = dt + relativedelta(months=1)
    return next_month.replace(day=1).strftime("%Y-%m-%d")

def end_date_calc(start: str, n_installments: int) -> str:
    dt = datetime.strptime(start, "%Y-%m-%d")
    end = dt + relativedelta(months=n_installments - 1)
    return end.replace(day=1).strftime("%Y-%m-%d")

def match_card(tarjeta_name: str, cards: dict) -> str | None:
    """Fuzzy match de nombre de tarjeta contra el diccionario de cards."""
    t = tarjeta_name.lower().strip()
    if not t:
        return None
    if t in cards:
        return cards[t]
    for k, v in cards.items():
        if k in t or t in k:
            return v
    for token in t.split():
        if len(token) >= 3:
            for k, v in cards.items():
                if token in k:
                    return v
    return None

# ─── Seed data ────────────────────────────────────────────────

SEED_CATEGORIES = [
    # Gastos
    {"name": "Casa",            "type": "expense", "color": "#6366f1", "emoji": "🏠"},
    {"name": "Supermercado",    "type": "expense", "color": "#10b981", "emoji": "🛒"},
    {"name": "Comida",          "type": "expense", "color": "#f97316", "emoji": "🍔"},
    {"name": "Transporte",      "type": "expense", "color": "#f59e0b", "emoji": "🚗"},
    {"name": "Uber",            "type": "expense", "color": "#1d4ed8", "emoji": "🚕"},
    {"name": "Salud",           "type": "expense", "color": "#ef4444", "emoji": "💊"},
    {"name": "Ejercicio",       "type": "expense", "color": "#84cc16", "emoji": "🏋️"},
    {"name": "Entretenimiento", "type": "expense", "color": "#8b5cf6", "emoji": "🎮"},
    {"name": "Tarjeta",         "type": "expense", "color": "#64748b", "emoji": "💳"},
    {"name": "Familia",         "type": "expense", "color": "#f472b6", "emoji": "👨‍👩‍👧"},
    {"name": "Sistemas",        "type": "expense", "color": "#06b6d4", "emoji": "💻"},
    {"name": "Inversión",       "type": "expense", "color": "#14b8a6", "emoji": "📈"},
    {"name": "Vacaciones",      "type": "expense", "color": "#fb923c", "emoji": "✈️"},
    {"name": "Otros",           "type": "expense", "color": "#6b7280", "emoji": "📦"},
    # Ingresos
    {"name": "Sueldo",          "type": "income",  "color": "#00e87a", "emoji": "💼"},
    {"name": "Regalo",          "type": "income",  "color": "#f472b6", "emoji": "🎁"},
    {"name": "Intereses",       "type": "income",  "color": "#0ea5e9", "emoji": "📈"},
    {"name": "Ajuste",          "type": "income",  "color": "#a78bfa", "emoji": "⚖️"},
    {"name": "Préstamo",        "type": "income",  "color": "#fb923c", "emoji": "🤝"},
    {"name": "Venta",           "type": "income",  "color": "#34d399", "emoji": "💰"},
]

SEED_CARDS = [
    {"name": "Visa Macro",   "bank": "Macro",   "color": "#00e87a"},
    {"name": "Amex Macro",   "bank": "Macro",   "color": "#1d4ed8"},
    {"name": "Naranja X",    "bank": "Naranja", "color": "#f59e0b"},
    {"name": "Mercado Pago", "bank": "MP",      "color": "#0ea5e9"},
]

def ensure_seed_data():
    """Inserta categorías y tarjetas base si no existen (idempotente)."""
    print("Asegurando categorias y tarjetas base...")

    # Categorias
    existing_cats = sb_get("categories", "select=name,type")
    existing_keys = {(c["name"].lower(), c["type"]) for c in existing_cats}
    missing_cats = [
        c for c in SEED_CATEGORIES
        if (c["name"].lower(), c["type"]) not in existing_keys
    ]
    if missing_cats:
        sb_post("categories", missing_cats)
        print(f"  + {len(missing_cats)} categorias insertadas")
    else:
        print(f"  = categorias ya existentes ({len(existing_cats)})")

    # Tarjetas
    existing_cards = sb_get("cards", "select=name")
    existing_names = {c["name"].lower() for c in existing_cards}
    missing_cards = [
        c for c in SEED_CARDS
        if c["name"].lower() not in existing_names
    ]
    if missing_cards:
        sb_post("cards", missing_cards)
        print(f"  + {len(missing_cards)} tarjetas insertadas")
    else:
        print(f"  = tarjetas ya existentes ({len(existing_cards)})")

# ─── Load reference data from Supabase ───────────────────────

def load_categories():
    cats = sb_get("categories", "select=id,name,type")
    return {
        "expense": {c["name"].lower(): c["id"] for c in cats if c["type"] == "expense"},
        "income":  {c["name"].lower(): c["id"] for c in cats if c["type"] == "income"},
    }

def load_cards():
    cards_data = sb_get("cards", "select=id,name")
    return {c["name"].lower(): c["id"] for c in cards_data}

# ─── Sheet parsers ────────────────────────────────────────────

def parse_gastos(ws, cats):
    """Hoja Gastos:
    col0=Fecha | col1=Categoría | col2=Detalle | col3=Monto ARS |
    col4=Monto U$D | col5=Monto U$D Google (#N/A) | col6=Dolar CCL (otro USD)
    CCL rate = monto_ars / monto_usd (calculado)
    """
    rows = []
    skipped = 0
    expense_cats = cats["expense"]

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        fecha = parse_date(row[0])
        if not fecha:
            skipped += 1
            continue

        cat_name = str(row[1] or "Otros").strip().lower()
        detalle = str(row[2] or "").strip()
        monto_ars = parse_float(row[3])
        monto_usd = parse_float_or_none(row[4] if len(row) > 4 else None)

        if monto_ars == 0:
            skipped += 1
            continue

        # CCL rate = ARS / USD (cuando el USD es válido)
        ccl = round(monto_ars / monto_usd, 2) if monto_usd and monto_usd > 0 else None

        cat_id = expense_cats.get(cat_name) or expense_cats.get("otros")
        if not cat_id:
            print(f"  ⚠️  Fila {i}: categoría gasto '{cat_name}' no encontrada → sin categoría")

        rows.append({
            "date": fecha,
            "category_id": cat_id,
            "detail": detalle or None,
            "amount": monto_ars,
            "currency": "ARS",
            "amount_ars": monto_ars,
            "amount_usd": monto_usd,
            "ccl_rate": ccl,
        })

    if skipped:
        print(f"  (saltadas {skipped} filas vacías/sin fecha en Gastos)")
    return rows

def parse_ingresos(ws, cats):
    """Hoja Ingresos: igual estructura que Gastos.
    CCL rate = monto_ars / monto_usd (calculado)
    """
    rows = []
    skipped = 0
    income_cats = cats["income"]

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        fecha = parse_date(row[0])
        if not fecha:
            skipped += 1
            continue

        cat_name = str(row[1] or "Sueldo").strip().lower()
        detalle = str(row[2] or "").strip()
        monto_ars = parse_float(row[3])
        monto_usd = parse_float_or_none(row[4] if len(row) > 4 else None)

        if monto_ars == 0:
            skipped += 1
            continue

        ccl = round(monto_ars / monto_usd, 2) if monto_usd and monto_usd > 0 else None

        cat_id = income_cats.get(cat_name) or income_cats.get("sueldo")
        if not cat_id:
            print(f"  ⚠️  Fila {i}: categoría ingreso '{cat_name}' no encontrada → sin categoría")

        rows.append({
            "date": fecha,
            "category_id": cat_id,
            "detail": detalle or None,
            "amount": monto_ars,
            "currency": "ARS",
            "amount_ars": monto_ars,
            "amount_usd": monto_usd,
            "ccl_rate": ccl,
        })

    if skipped:
        print(f"  (saltadas {skipped} filas vacías/sin fecha en Ingresos)")
    return rows

def parse_datos(ws, cards):
    """Hoja Datos:
    col0=Fecha compra | col1=Descripción | col2=Monto total | col3=Monto Abonado por mi |
    col4=Cuotas | col5=Tarjeta | col6=Cuota mensual | col7=Cuota mensual Abonada |
    col8=Fecha pago inicial | col9=Fecha pago final | col10=USD (#N/A)
    """
    rows = []
    skipped = 0

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        fecha_compra = parse_date(row[0])
        if not fecha_compra:
            skipped += 1
            continue

        descripcion = str(row[1] or "").strip()
        if not descripcion:
            skipped += 1
            continue

        monto_total = parse_float(row[2])
        if monto_total == 0:
            skipped += 1
            continue

        monto_abonado = parse_float(row[3]) or monto_total
        cuotas = max(1, int(parse_float(row[4]) or 1))
        tarjeta_name = str(row[5] or "").strip()

        card_id = match_card(tarjeta_name, cards)
        if not card_id and tarjeta_name:
            print(f"  ⚠️  Fila {i}: tarjeta '{tarjeta_name}' no encontrada → sin tarjeta")

        # Usar fechas ya calculadas en el Excel (cols 8 y 9) si existen
        start = parse_date(row[8]) if len(row) > 8 and row[8] else None
        e_date = parse_date(row[9]) if len(row) > 9 and row[9] else None
        if not start:
            start = first_day_next_month(fecha_compra)
        if not e_date:
            e_date = end_date_calc(start, cuotas)

        # Cuotas pagadas hasta hoy (incluyendo el mes actual si ya pasó el día 1)
        today = datetime.today()
        start_dt = datetime.strptime(start, "%Y-%m-%d")
        # +1 porque el mes de inicio cuenta como cuota pagada una vez que llegó su día
        months_elapsed = (today.year - start_dt.year) * 12 + (today.month - start_dt.month) + 1
        paid = max(0, min(months_elapsed, cuotas))

        rows.append({
            "description": descripcion,
            "card_id": card_id,
            "total_amount": monto_total,
            "paid_amount": monto_abonado,
            "currency": "ARS",
            "total_installments": cuotas,
            "paid_installments": paid,
            "start_date": start,
            "end_date": e_date,
            "counts_towards_balance": True,
        })

    if skipped:
        print(f"  (saltadas {skipped} filas sin datos en Datos)")
    return rows

def parse_exchange_rates(gastos_ws, ingresos_ws):
    """Extrae cotizaciones CCL únicas.
    CCL rate = monto_ars (col3) / monto_usd (col4), una por fecha.
    Toma la última tasa vista para cada fecha.
    """
    rates: dict[str, float] = {}
    for ws in [gastos_ws, ingresos_ws]:
        if ws is None:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            fecha = parse_date(row[0])
            monto_ars = parse_float(row[3]) if len(row) > 3 else 0
            monto_usd = parse_float_or_none(row[4] if len(row) > 4 else None)
            if fecha and monto_ars > 0 and monto_usd and monto_usd > 0:
                rates[fecha] = round(monto_ars / monto_usd, 2)
    return [{"date": d, "ccl_rate": r} for d, r in sorted(rates.items())]

# ─── Main ─────────────────────────────────────────────────────

def batch(lst, size=100):
    for i in range(0, len(lst), size):
        yield lst[i:i+size]

def show_preview(rows, label, n=3):
    if not rows:
        return
    print(f"\n   Primeras {min(n, len(rows))} filas de {label}:")
    for r in rows[:n]:
        print(f"     {r}")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="Ruta al archivo .xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Solo muestra qué haría, sin insertar")
    parser.add_argument("--show-sheets", action="store_true", help="Muestra el nombre y primeras filas de cada hoja")
    parser.add_argument("--skip-rates", action="store_true", help="No importar cotizaciones CCL")
    args = parser.parse_args()

    xlsx_path = Path(args.file)
    if not xlsx_path.exists():
        sys.exit(f"Archivo no encontrado: {xlsx_path}")

    print(f"📂 Leyendo {xlsx_path.name}...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    print(f"   Hojas encontradas: {wb.sheetnames}")

    if args.show_sheets:
        for name in wb.sheetnames:
            ws = wb[name]
            print(f"\n── Hoja: {name} ──")
            for j, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True)):
                print(f"  Fila {j+1}: {list(row)}")
        return

    sheets = {name.lower(): wb[name] for name in wb.sheetnames}

    ensure_seed_data()
    print("📡 Cargando categorías y tarjetas de Supabase...")
    cats = load_categories()
    cards = load_cards()
    print(f"   Categorías expense: {list(cats['expense'].keys())}")
    print(f"   Categorías income:  {list(cats['income'].keys())}")
    print(f"   Tarjetas:           {list(cards.keys())}")

    gastos_sheet   = sheets.get("gastos")   or sheets.get("gasto")
    ingresos_sheet = sheets.get("ingresos") or sheets.get("ingreso")
    datos_sheet    = sheets.get("datos")    or sheets.get("cuotas")

    if not gastos_sheet:
        print(f"⚠️  No se encontró hoja 'Gastos' (disponibles: {list(sheets.keys())})")
    if not ingresos_sheet:
        print(f"⚠️  No se encontró hoja 'Ingresos' (disponibles: {list(sheets.keys())})")
    if not datos_sheet:
        print(f"⚠️  No se encontró hoja 'Datos' (disponibles: {list(sheets.keys())})")

    print("\n📊 Parseando hojas...")
    gastos_rows   = parse_gastos(gastos_sheet, cats)            if gastos_sheet   else []
    ingresos_rows = parse_ingresos(ingresos_sheet, cats)         if ingresos_sheet else []
    datos_rows    = parse_datos(datos_sheet, cards)              if datos_sheet    else []
    rates_rows    = parse_exchange_rates(gastos_sheet, ingresos_sheet) if not args.skip_rates else []

    print(f"\n📋 Resumen:")
    print(f"   Gastos:         {len(gastos_rows)}")
    print(f"   Ingresos:       {len(ingresos_rows)}")
    print(f"   Cuotas:         {len(datos_rows)}")
    print(f"   Cotizaciones:   {len(rates_rows)}")

    if args.dry_run:
        show_preview(gastos_rows,   "Gastos")
        show_preview(ingresos_rows, "Ingresos")
        show_preview(datos_rows,    "Cuotas")
        show_preview(rates_rows,    "Cotizaciones")
        print("\n⚠️  Dry run — no se insertó nada.")
        return

    confirm = input("\n¿Insertar todo en Supabase? (s/n): ").strip().lower()
    if confirm != "s":
        print("Cancelado.")
        return

    # Insert cotizaciones CCL (ignorar conflictos de fecha única)
    if rates_rows:
        print(f"\n⬆️  Insertando {len(rates_rows)} cotizaciones CCL...")
        headers_upsert = {**HEADERS, "Prefer": "resolution=ignore-duplicates,return=representation"}
        r = requests.post(f"{SUPABASE_URL}/rest/v1/exchange_rates", headers=headers_upsert, json=rates_rows)
        if r.ok:
            inserted = len(r.json()) if r.text else 0
            print(f"   ✓ {inserted} insertadas ({len(rates_rows) - inserted} ya existían)")
        else:
            print(f"   ERROR: {r.text[:200]}")

    # Insert gastos
    if gastos_rows:
        print(f"\n⬆️  Insertando {len(gastos_rows)} gastos...")
        inserted = 0
        for chunk in batch(gastos_rows):
            result = sb_post("expenses", chunk)
            inserted += len(result)
        print(f"   ✓ {inserted} insertados")

    # Insert ingresos
    if ingresos_rows:
        print(f"\n⬆️  Insertando {len(ingresos_rows)} ingresos...")
        inserted = 0
        for chunk in batch(ingresos_rows):
            result = sb_post("incomes", chunk)
            inserted += len(result)
        print(f"   ✓ {inserted} insertados")

    # Insert cuotas
    if datos_rows:
        print(f"\n⬆️  Insertando {len(datos_rows)} compras en cuotas...")
        inserted = 0
        for chunk in batch(datos_rows):
            result = sb_post("installment_purchases", chunk)
            inserted += len(result)
        print(f"   ✓ {inserted} insertadas")

    print("\n✅ Migración completa.")

if __name__ == "__main__":
    main()
